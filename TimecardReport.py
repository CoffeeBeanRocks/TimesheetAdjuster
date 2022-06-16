# @Author: Ethan Meyers
# Email: ewm1230@gmail.com
# Phone#: 847-212-2264
# @Date: 06/13/2022

# This script takes a path to an Excel file
# and outputs an Excel sheet that filters all names
# that are on the 1099 driver list

import pandas as pd
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import numbers, PatternFill
from datetime import datetime, timedelta


def get1099Drivrs():
    logins = [
        'cculp',
        'jlorden',
        'fherron',
        'ekotlarz',
        'ssetterlund',
        'kjohnson',
        'rgrover',
        'mmartin',
        'gkozlowski',
        'zkaczmarczyk',
        'bmihaylov',
        'kslupek',
        'ibuzinskis',
        'PWALAWSKI',
        'VDIMITROV',
        'mrudzinski',
        'aszymanski',
        'awiech',
        'mwiechetek',
        'awolak',
        'mzareba',
        'shicks',
        'sutterback',
        'ffrench',
        'jgrzesiak',
        'bgal',
        'jmitchell',
        'kpopek',
        'mlooney',
        'sgorczyca',
        'jhaynie',
        'epetrov',
        'lbronikowski',
        'rgranados',
        'kpodstawka',
        'rbanas',
        'khadera',
        'rpettis',
        'aplecki',
        'kwasowicz',
        'kbryja',
        'jsadkowski',
        'jlukacs',
        'miwaniec',
        'tcachro',
        'rsingh',
        'dgornikowski',
        'michaelj',
        'lbalinski',
        'rkokot',
        'dzajac',
        'sspear',
        'tbrown',
        'sbliznakov',
        'jfoxx',
        'skahlon',
        'bmontgomery',
        'jhernandez',
        'dbooker',
        'tvarela',
        'rwyrick',
        'rwade',
        'hhernandez',
        'jrzepka',
        'dfidowski',
        'nlewis',
        'mholder',
        'jchavez',
        'brobbins',
        'smartinez',
        'rpetraitis',
        'saddison',
        'tthompson',
        'eagbenyadzi',
        'iasenov',
        'ccortes',
        'omara',
        'jramirez',
        'aradon',
        'apatino',
        'awilliams',
        'psoja',
        'ccardona'
    ]
    drivers = pd.DataFrame(logins, columns=['1099 Drivers'])
    return drivers

def deleteRows(FilePath):
    # Gets data from Excel sheet and removes elements that are in the 1099 drivers list
    df = pd.read_excel(FilePath, sheet_name='Duty Time', header=8)
    # df2 = pd.read_excel(FilePath, sheet_name='Sheet1', header=0)
    df2 = get1099Drivrs()
    df = df[~df['Login'].isin(df2['1099 Drivers'])]

    # Removes empty columns
    df.replace("", "NaN", inplace=True)
    df.dropna(subset=['Login'], inplace=True)

    # Removes mystery total column
    del df['Total']
    del df['Unnamed: 0']

    # with pd.option_context('display.max_rows', None, 'display.max_columns', None):
    #     print(df)

    # Print DF
    writer = pd.ExcelWriter(FilePath, engine='openpyxl')
    df.to_excel(writer, sheet_name='Output')
    writer.save()


def formatCols(FilePath):
    workbook = openpyxl.load_workbook(FilePath)
    worksheet = workbook['Output']

    # Using openpyxl to format the new Excel sheet

    # Format Source Code: https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
    for cell in worksheet['D']:
        cell.number_format = numbers.FORMAT_DATE_XLSX14
    for cell in worksheet['G']:
        cell.number_format = numbers.FORMAT_DATE_XLSX14

    for cell in worksheet['E']:
        cell.number_format = numbers.FORMAT_DATE_TIME2
    for cell in worksheet['H']:
        cell.number_format = numbers.FORMAT_DATE_TIME2

    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 20


    # Alternates filling each row based off of a new username
    yellowFill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
    currentName = worksheet['B2'].value
    colIndex = 2
    fill = False
    for i in range(2, worksheet.max_row+1):
        if worksheet.cell(row=i, column=colIndex).value == currentName and fill:
            for j in range(colIndex, worksheet.max_column+1):
                worksheet.cell(row=i, column=j).fill = yellowFill
        else:
            if worksheet.cell(row=i, column=colIndex).value != currentName:
                fill = not fill
                currentName = worksheet.cell(row=i, column=colIndex).value
                if fill:
                    for j in range(colIndex, worksheet.max_column+1):
                        worksheet.cell(row=i, column=j).fill = yellowFill

    redFill = PatternFill(start_color='00FC0303', end_color='00FC0303', fill_type='solid')
    orangeFill = PatternFill(start_color='00FAAC02', end_color='00FAAC02', fill_type='solid')
    currentName = worksheet['B2'].value
    startIndex = 2
    fill = False
    for i in range(2, worksheet.max_row + 2):
        if worksheet.cell(row=i, column=colIndex).value != currentName:
            currentName = worksheet.cell(row=i, column=colIndex).value
            total = timedelta(hours=0)
            for j in range(startIndex, i):
                s1 = str(worksheet.cell(row=j, column=5).value)
                startTime = datetime.strptime(s1, '%Y-%m-%d %H:%M:%S')
                s2 = str(worksheet.cell(row=j, column=8).value)
                endTime = datetime.strptime(s2, '%Y-%m-%d %H:%M:%S')
                total = total + (endTime-startTime)
            worksheet.cell(row=i - 1, column=9).value = total
            if fill:
                worksheet.cell(row=i - 1, column=9).fill = orangeFill
            else:
                worksheet.cell(row=i - 1, column=9).fill = redFill
            fill = not fill
            startIndex = i

    workbook.save(FilePath)

if __name__ == '__main__':
    print("Instructions: ", "1) Find the timecard sheet in file-explorer or on the desktop", "2) Right click the file and select, \"Copy\"", "3) Right click in file-explorer or on the desktop and select, \"Paste\"", "4) Right click the copied file and select, \"Copy as path\"", sep='\n')
    print("5) You've just copied the file-path to your clipboard, press \"CTRL V\" and paste the path below. Then press enter: ")
    path = input()
    if '"' in path:
        path = path.replace('"', '')
    try:
        print("Formatting File, Please Wait!")
        deleteRows(path)
        formatCols(path)
        print("Task Completed!")
    except Exception as e:
        if "Errno 13" in str(e):
            print("Please close the file and try again!")
        else:
            print("An unknown error occurred:", e)
    input("\nPress enter to finish: ")


