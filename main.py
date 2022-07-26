# @Author: Ethan Meyers
# @Email: ewm1230@gmail.com
# @Phone: 847-212-2264
# @Date: 06/13/2022

# This script takes a path to an Excel file
# and outputs an Excel sheet that filters all names
# that are on the 1099 driver list


import datetime
from datetime import datetime, timedelta, time
import os
import sys
import pandas as pd
from os.path import exists
import openpyxl
from openpyxl.styles import numbers, PatternFill


# @Var dir_Path: Path to the folder containing the necessary files for this program
# @Var w2Path: Path to the file containing the names of the W2 drivers
# @Description: Class for necessary global variables
class Data:
    dir_path = '%s\\HOSFilter\\' % os.environ['APPDATA']
    w2Path = '%sDrivers.xlsx' % dir_path


# @Description: Verifies that the proper directories and files are in place
def loadData():
    if not os.path.exists(Data.dir_path):
        os.makedirs(Data.dir_path)

    if not exists(Data.w2Path):
        print('List of W2 Drivers not found, please enter new list!')
        copyXLSX(input("Enter new W2 list: "))


# @Param filePath: File path to the new Drivers.xlsx file provided by the user
# @Description: Fills the drivers.xlsx file with names given from a different file
def copyXLSX(filePath: str):
    if '"' in filePath:
        filePath = filePath.replace('"', '')
    df = pd.read_excel(filePath, sheet_name='Sheet1', header=0)
    writer = pd.ExcelWriter(Data.w2Path, engine='openpyxl')
    df.columns = ["W2 Drivers"]
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()


# @Return: Dataframe of all the names of the W2 Drivers
# @Description: Finds the file of Excel drivers and returns a dataframe of all their names
def getW2() -> pd.DataFrame:
    if not exists(Data.w2Path):
        loadData()

    drivers = pd.read_excel(Data.w2Path, sheet_name='Sheet1', header=0)
    drivers['W2 Drivers'] = drivers['W2 Drivers'].str.lower()
    return drivers


# @Param td: A timedelta object
# @Return str: The string value of the total hours:minutes:seconds represented by the timedelta object
# @Description: Takes a time delta object and converts it into the HH:MM:SS format
def formatTimedelta(td: timedelta) -> str:
    minutes, seconds = divmod(td.seconds + td.days * 86400, 60)
    hours, minutes = divmod(minutes, 60)
    return '{:d}:{:02d}:{:02d}'.format(hours, minutes, seconds)


# @Param filePath: File path to the user's input file
# @Description: Overwrites input file and outputs all the names and their clock in/out if they're a W2 Driver
def deleteRows(filePath: str):
    # Filters all the drivers on the input file
    df = pd.read_excel(filePath, sheet_name='Duty Time', header=8, index_col=False)
    df2 = getW2()
    df = df[df['Login'].str.lower().isin(df2['W2 Drivers'])]

    # Removes empty columns
    df.replace("", "NaN", inplace=True)
    df.dropna(subset=['Login'], inplace=True)

    # Removes total column and unnamed column from dataframe
    del df['Total']
    del df['Unnamed: 0']

    # Calculates the total hours clocked in for each person
    total = timedelta(hours = 0)
    name = df.iloc[0]['Login']
    totals = []
    for i in range(0, len(df.index)):
        if name != df.iloc[i]['Login']:
            totals.append(formatTimedelta(total))
            name = df.iloc[i]['Login']
            total = timedelta(hours = 0)
        elif i > 0:
            totals.append("")
        dtEnd = df.iloc[i]['Shift End Time']
        dtStart = df.iloc[i]['Shift Start Time']
        total += (dtEnd - dtStart)
    totals.append(formatTimedelta(total))

    # Overwrite input file with the filtered data
    df['Total Hours'] = totals
    writer = pd.ExcelWriter(filePath, engine='openpyxl')
    df.to_excel(writer, sheet_name='Output', index=False)
    writer.save()


# @Param filePath: File path to the output file
# Description: Formats the output data into a more readable format
def formatCols(filePath: str):
    # Setup for formatting output file
    workbook = openpyxl.load_workbook(filePath)
    worksheet = workbook['Output']

    # Changes the column data format
    # Format Source Code: https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
    for cell in worksheet['C']:
        cell.number_format = numbers.FORMAT_DATE_XLSX14
    for cell in worksheet['F']:
        cell.number_format = numbers.FORMAT_DATE_XLSX14

    for cell in worksheet['D']:
        cell.number_format = numbers.FORMAT_DATE_TIME2
    for cell in worksheet['G']:
        cell.number_format = numbers.FORMAT_DATE_TIME2

    # Changes the column's width
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 20

    # Alternates filling each row based off of a new username
    yellowFill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
    currentName = worksheet['A2'].value
    colIndex = 1
    fill = False
    for i in range(colIndex, worksheet.max_row+1):
        if worksheet.cell(row=i, column=colIndex).value == currentName and fill:
            for j in range(colIndex, worksheet.max_column+1):
                worksheet.cell(row=i, column=j).fill = yellowFill
        else:
            if worksheet.cell(row=i, column=colIndex).value != currentName:
                fill = not fill
                currentName = worksheet.cell(row=i, column=colIndex).value
                if fill:
                    for j in range(colIndex, worksheet.max_column+1):
                        worksheet.cell(row=i, column=j).fill = yellowFill

    # Save updates
    workbook.save(filePath)


# @Description: Main runner for the program
if __name__ == '__main__':
    loadData()
    print("Instructions: ", "Before running the program make sure the relevant Excel file is closed", "1) Find the timecard sheet in file-explorer or on the desktop", "2) Right click the file and select, \"Copy\"", "3) Right click in file-explorer or on the desktop and select, \"Paste\"", "4) Hold shift then right click the copied file and select, \"Copy as path\"", sep='\n')
    print("5) You've just copied the file-path to your clipboard, press \"CTRL V\" and paste the path below. Then press enter")
    path = input("Paste on this line here: ")
    if '"' in path:
        path = path.replace('"', '')
    elif "-v" == path:
        print(getW2())
        input("\nPress enter to finish: ")
        sys.exit("Finished!")
    elif "-c" == path:
        print('W2 list amendment mode, please enter new list!')
        print('All data must be on a sheet titled "Sheet1" and have a header at "A0" followed by the data')
        copyXLSX(input("Enter new W2 list: "))
        input("\nPress enter to finish: ")
        sys.exit("Finished!")
    try:
        print("Formatting File, Please Wait!")
        deleteRows(path)
        formatCols(path)
        print("Task Completed!")
    except Exception as e:
        if "Errno 13" in str(e):
            print("Please close the file and try again!")
        else:
            print("An unknown error occurred:", e)
    input("\nPress enter to finish: ")